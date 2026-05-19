#!/usr/bin/env python3
"""
DevCloud Sprint Tasks → Excel
Usage:
    opencli devcloud sprint-tasks --url "<url>" --format csv | python excel/export_to_excel.py
    python excel/export_to_excel.py --input tasks.csv --output excel/output.xlsx
"""
import sys
import csv
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def style_header_row(ws, col_count):
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def style_data_rows(ws, row_count, col_count):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    for row in range(2, row_count + 2):
        fill = alt_fill if row % 2 == 0 else PatternFill()
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def auto_column_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)


def _clean(val):
    """Strip lone surrogate characters that openpyxl/lxml cannot encode."""
    if isinstance(val, str):
        return val.encode('utf-8', 'surrogatepass').decode('utf-8', 'replace')
    return val


def export(reader, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sprint Tasks"
    ws.freeze_panes = "A2"

    rows = list(reader)
    if not rows:
        print("ERROR: No data to export (empty input).", file=sys.stderr)
        sys.exit(1)

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([_clean(row.get(h, "")) for h in headers])

    style_header_row(ws, len(headers))
    style_data_rows(ws, len(rows), len(headers))
    auto_column_width(ws)
    ws.row_dimensions[1].height = 22

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Exported {len(rows)} rows → {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Export DevCloud sprint tasks to Excel")
    parser.add_argument("--input", "-i", help="CSV input file (default: stdin)")
    parser.add_argument(
        "--output", "-o",
        help="Output .xlsx path (default: excel/sprint_tasks_<timestamp>.xlsx)"
    )
    args = parser.parse_args()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_out = Path(__file__).parent / f"sprint_tasks_{timestamp}.xlsx"
    output_path = Path(args.output) if args.output else default_out

    if args.input:
        with open(args.input, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            export(reader, output_path)
    else:
        # Read from stdin
        text = sys.stdin.read()
        if not text.strip():
            print("ERROR: No input received on stdin.", file=sys.stderr)
            sys.exit(1)
        import io
        reader = csv.DictReader(io.StringIO(text))
        export(reader, output_path)


if __name__ == "__main__":
    main()
