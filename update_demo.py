#!/usr/bin/env python3
"""
update_demo.py - Extract DevCloud task/bug details and append to demo.xlsx.

Routing (status checked first):
  status in {新建, 进行中}  -> 遗留问题 sheet
  status else + Bug         -> 修复缺陷 sheet
  status else + other       -> 新增功能 sheet

新增功能 / 遗留问题 columns (A-I):
  ID | Title | EndDate | Status | Assignee | EstStart | EstEnd | Priority | Created

修复缺陷 columns (A-K, skip F/G):
  ID | Title | Status | Assignee | Priority | - | - | EndDate | EstStart | EstEnd | Created

Usage:
    python update_demo.py <url1> [url2 ...]
    python update_demo.py --file urls.txt
"""

import subprocess
import json
import re
import sys
import os
from datetime import datetime, timezone

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SESSION = "devcloud-tasks"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR  = os.path.join(SCRIPT_DIR, "excel")


def _resolve_xlsx(explicit=None):
    """Return path to the target xlsx. Auto-discover if not specified."""
    if explicit:
        return explicit
    found = [f for f in os.listdir(EXCEL_DIR) if f.lower().endswith(".xlsx")]
    if len(found) == 1:
        return os.path.join(EXCEL_DIR, found[0])
    if len(found) == 0:
        print(f"ERROR: no .xlsx files found in {EXCEL_DIR}", file=sys.stderr)
        sys.exit(1)
    print(f"Multiple xlsx files found — use --excel to specify one:", file=sys.stderr)
    for f in found:
        print(f"  {f}", file=sys.stderr)
    sys.exit(1)

SHEET_FEATURE = "\u65b0\u589e\u529f\u80fd "  # 新增功能 (trailing space in xlsx)
SHEET_BUG     = "\u4fee\u590d\u7f3a\u9677"   # 修复缺陷
SHEET_LEGACY  = "\u9057\u7559\u95ee\u9898"   # 遗留问题
SHEET_VERSION = "\u7248\u672c\u4fe1\u606f"   # 版本信息

# Statuses that route to the legacy (遗留问题) sheet regardless of type
LEGACY_STATUSES = {"\u65b0\u5efa", "\u8fdb\u884c\u4e2d"}  # 新建, 进行中

# JavaScript: synchronous XHR to the same-origin getShow API.
# X-Requested-With header is required - without it the SPA returns its HTML shell.
EXTRACT_JS = (
    "(function(){"
    r"var m=location.href.match(/\/scrum\/([a-f0-9]+)\/.*\/detail\/(\d+)/);"
    "if(!m)return JSON.stringify({error:'url_mismatch',href:location.href});"
    "var pid=m[1],iid=m[2];"
    "var xhr=new XMLHttpRequest();"
    "xhr.open('GET','/projectman/openapi/v1/scrum/getShow?projectUUId='+pid+'&issueId='+iid+'&include=children,parent&limit=10&offset=0',false);"
    "xhr.setRequestHeader('X-Requested-With','XMLHttpRequest');"
    "xhr.setRequestHeader('Accept','application/json');"
    "xhr.send();"
    "var d;"
    "try{d=JSON.parse(xhr.responseText);}catch(e){return JSON.stringify({error:'parse_error',body:xhr.responseText.substring(0,200)});}"
    "if(!d||!d.result||!d.result.issue)return JSON.stringify({error:'unexpected_shape',keys:Object.keys(d||{}).join(',')});"
    "var iss=d.result.issue;"
    "return JSON.stringify({"
    "id:iss.id,"
    "subject:iss.subject||'',"
    "status:iss.status&&iss.status.name||'',"
    "assignee:iss.assigned_to&&iss.assigned_to.firstName||'',"
    "priority:iss.priority&&iss.priority.name||'',"
    "type:iss.tracker&&iss.tracker.name||'',"
    "due_date:iss.due_date||'',"
    "start_date:iss.start_date||'',"
    "created_on:iss.created_on||''"
    "});"
    "})()"
)


def run_opencli(*args):
    def _q(s):
        s = str(s).replace('"', '\\"')
        return f'"{s}"'
    cmd = " ".join(_q(a) for a in ["opencli", "browser", SESSION] + list(args))
    r = subprocess.run(cmd, shell=True, capture_output=True,
                       text=True, encoding="utf-8", errors="replace")
    return r.stdout.strip(), r.stderr.strip(), r.returncode


def _parse_date(val):
    """Convert ms-epoch string or ISO datetime to YYYY-MM-DD."""
    if not val:
        return ""
    s = str(val).strip()
    if re.fullmatch(r"\d{10,}", s):
        try:
            dt = datetime.fromtimestamp(int(s) / 1000, tz=timezone.utc)
            return dt.strftime("%Y-%m-%d")
        except (ValueError, OverflowError):
            return s
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    return m.group(1) if m else s


def extract_task_number(url):
    m = re.search(r"/detail/(\d+)", url) or re.search(r"/(\d{7,})(?:/|$)", url)
    return m.group(1) if m else ""


def fetch_task(url):
    task_num = extract_task_number(url)
    print(f"\n[#{task_num}] opening...", flush=True)

    _, err, code = run_opencli("open", url)
    if code != 0:
        print(f"  open error: {err}", flush=True)
        return None

    run_opencli("wait", "selector",
        "[class*='detail'],[class*='workitem'],[class*='work-item'],[class*='issue'],main",
        "--timeout", "10000")
    run_opencli("wait", "time", "2")

    eval_out, eval_err, eval_code = run_opencli("eval", EXTRACT_JS)
    if eval_code != 0 or not eval_out:
        print(f"  eval error: {eval_err or eval_out}", flush=True)
        return None

    try:
        data = json.loads(eval_out)
    except json.JSONDecodeError:
        print(f"  parse error: {eval_out[:120]}", flush=True)
        return None

    if "error" in data:
        print(f"  API error: {data}", flush=True)
        return None

    r = {
        "id":         str(data.get("id", task_num)),
        "url":        url,
        "subject":    data.get("subject", ""),
        "status":     data.get("status", ""),
        "assignee":   data.get("assignee", ""),
        "priority":   data.get("priority", "").strip(),
        "type":       data.get("type", ""),
        "due_date":   _parse_date(data.get("due_date", "")),
        "start_date": _parse_date(data.get("start_date", "")),
        "created_on": _parse_date(data.get("created_on", "")),
    }
    print(f"  [{r['type']}] {r['subject'][:50]!r}  status={r['status']!r}  priority={r['priority']!r}", flush=True)
    return r


# JS: try multiple GitLab selectors to find a 7-40 char hex commit hash.
_COMMIT_JS = (
    "(function(){"
    "var sels=['.commit-sha','[data-testid=\"commit-sha-group\"] a','.label-monospace','a[href*=\"/commit/\"]'];"
    "for(var i=0;i<sels.length;i++){"
    "var el=document.querySelector(sels[i]);"
    "if(el){var t=el.textContent.trim();if(/^[0-9a-f]{7,40}$/.test(t))return t;}"
    "}"
    "var btns=document.querySelectorAll('[data-clipboard-text]');"
    "for(var j=0;j<btns.length;j++){"
    "var v=btns[j].getAttribute('data-clipboard-text');"
    "if(v&&/^[0-9a-f]{7,40}$/.test(v.trim()))return v.trim();"
    "}"
    "return '';"
    "})()"
)


def fetch_commit(build_url):
    """Navigate to the GitLab tree page and return the latest commit hash."""
    print(f"\n[version] fetching commit from {build_url} ...", flush=True)
    run_opencli("open", build_url)
    run_opencli("wait", "selector", ".commit-sha,.label-monospace,[data-clipboard-text]", "--timeout", "10000")
    run_opencli("wait", "time", "1")
    out, _, code = run_opencli("eval", _COMMIT_JS)
    if code == 0 and out and re.fullmatch(r"[0-9a-f]{7,40}", out.strip()):
        return out.strip()
    print(f"  [version] could not extract commit (got: {out!r})", flush=True)
    return None


def _id_in_sheet(ws, task_id, id_col=1):
    for row in ws.iter_rows(min_row=2, max_col=id_col, values_only=True):
        if str(row[0]) == str(task_id):
            return True
    return False


def _next_empty_row(ws):
    for i in range(2, ws.max_row + 2):
        if all(ws.cell(row=i, column=c).value is None for c in range(1, 5)):
            return i
    return ws.max_row + 1


def update_excel(rows, demo_path):
    if not os.path.exists(demo_path):
        print(f"ERROR: {demo_path} not found.", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(demo_path)
    added_feat = added_bug = added_legacy = skipped = 0

    for r in rows:
        if r is None:
            continue
        is_legacy = r["status"] in LEGACY_STATUSES
        is_bug    = r["type"].lower() in ("bug", "\u7f3a\u9677")

        if is_legacy:
            ws = wb[SHEET_LEGACY]
        elif is_bug:
            ws = wb[SHEET_BUG]
        else:
            ws = wb[SHEET_FEATURE]

        if _id_in_sheet(ws, r["id"]):
            print(f"  [skip] #{r['id']} already in '{ws.title}'", flush=True)
            skipped += 1
            continue

        row_i = _next_empty_row(ws)

        if is_bug and not is_legacy:
            # 修复缺陷: A编号 B标题 C状态 D处理人 E优先级 (F完成度 G类型 skip) H结束时间 I预计开始 J预计结束 K创建时间
            c = ws.cell(row_i, 1); c.value = r["id"]; c.hyperlink = r["url"]
            ws.cell(row_i, 2).value  = r["subject"]
            ws.cell(row_i, 3).value  = r["status"]
            ws.cell(row_i, 4).value  = r["assignee"]
            ws.cell(row_i, 5).value  = r["priority"]
            ws.cell(row_i, 8).value  = r["due_date"]
            ws.cell(row_i, 9).value  = r["start_date"]
            ws.cell(row_i, 10).value = r["due_date"]
            ws.cell(row_i, 11).value = r["created_on"]
            added_bug += 1
        else:
            # 新增功能 / 遗留问题: A编号 B标题 C结束时间 D状态 E处理人 F预计开始 G预计结束 H优先级 I创建时间
            c = ws.cell(row_i, 1); c.value = r["id"]; c.hyperlink = r["url"]
            ws.cell(row_i, 2).value = r["subject"]
            ws.cell(row_i, 3).value = r["due_date"]
            ws.cell(row_i, 4).value = r["status"]
            ws.cell(row_i, 5).value = r["assignee"]
            ws.cell(row_i, 6).value = r["start_date"]
            ws.cell(row_i, 7).value = r["due_date"]
            ws.cell(row_i, 8).value = r["priority"]
            ws.cell(row_i, 9).value = r["created_on"]
            if is_legacy:
                added_legacy += 1
            else:
                added_feat += 1

        print(f"  written #{r['id']} -> '{ws.title}' row {row_i}", flush=True)

    # Update 版本信息: read build URL from B3, fetch latest commit, write date+commit
    try:
        vs = wb[SHEET_VERSION]
        build_url = vs.cell(3, 2).value
        if build_url:
            commit = fetch_commit(str(build_url).strip())
            today = datetime.now().strftime("%Y%m%d")
            vs.cell(2, 2).value = today
            if commit:
                vs.cell(4, 2).value = commit
            print(f"  [version] 构建日期={today}  commit={commit or '(unchanged)'}", flush=True)
    except KeyError:
        print(f"  [version] sheet '{SHEET_VERSION}' not found, skipping", flush=True)

    wb.save(demo_path)
    print(f"\nSaved: {demo_path}", flush=True)

    # Sync to Word release-notes document if update_word.py is present
    try:
        sys.path.insert(0, SCRIPT_DIR)
        from update_word import update_word as _uw
        # Pass explicit excel path so we don't re-resolve
        _uw(excel_path=demo_path)
    except ImportError:
        pass  # update_word.py optional
    except Exception as e:
        print(f"  [word] WARNING: {e}", flush=True)

    print(f"  {SHEET_FEATURE}: +{added_feat}  {SHEET_BUG}: +{added_bug}  {SHEET_LEGACY}: +{added_legacy}  skipped: {skipped}", flush=True)


def main():
    args = sys.argv[1:]
    urls = []

    # --excel <path>: explicit xlsx target
    demo_path = None
    if "--excel" in args:
        idx = args.index("--excel")
        if idx + 1 >= len(args):
            print("Error: --excel requires a path argument", file=sys.stderr)
            sys.exit(1)
        demo_path = args[idx + 1]
        args = args[:idx] + args[idx + 2:]
    demo_path = _resolve_xlsx(demo_path)

    if "--file" in args:
        idx = args.index("--file")
        if idx + 1 >= len(args):
            print("Error: --file requires a path argument", file=sys.stderr)
            sys.exit(1)
        with open(args[idx + 1], "r", encoding="utf-8") as f:
            urls = [ln.strip() for ln in f if ln.strip() and not ln.startswith("#")]
    else:
        urls = [a for a in args if a.startswith("http")]

    if not urls:
        print(__doc__)
        sys.exit(1)

    print(f"Fetching {len(urls)} task(s)...", flush=True)
    rows = [fetch_task(u) for u in urls]
    update_excel([r for r in rows if r is not None], demo_path)


if __name__ == "__main__":
    main()